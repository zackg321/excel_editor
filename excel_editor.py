import openpyxl
from openpyxl import Workbook, load_workbook

import datetime
from datetime import datetime

#A Python program to insert data into TC excel workbook. 

#Enter in information including stage, date, number of explants
def get_input():
    wb_name = input("Enter customer name: ")
    strain =  input("Enter strain name: ")
    stage = input("Enter the stage: ")
    explants = int(input("Enter number of explants: "))
    date = (input("Enter the date as MM/DD/YYYY: "))
    date_obj = datetime.strptime(date, "%m/%d/%Y")
    date = date_obj
    data_list = [stage, date, explants, wb_name, strain]
    return data_list

def get_column(stage):
    if stage == str("1"):
        stage = 4
    elif stage == str("2"):
        stage = 8
    elif stage == "MS":
        stage = 16
    elif stage == str("3"):
        stage = 12
    else:
        print("Stage not found")
    return int(stage)

def get_empty_row(stage, ws):
    column = int(get_column(stage) + 1)
    rows = ws.iter_rows(min_row = 3, max_row = ws.max_row + 1, min_col = column, max_col = column)
    #iterate through rows to find empty row
    for a in rows:
        if a[0].value == None:
            cell_row = a[0].row
            break
    return (cell_row, column - 1)

def enter_data(data, ws):
    coordinates = get_empty_row(data[0], ws)

    ws.cell(row = coordinates[0], column = coordinates[1]).value = data[1]

    ws.cell(row = coordinates[0], column = coordinates[1] + 1).value = data[2]

def run():
    more_data = True
    while more_data:
        data = get_input()
        #Load workbook
        wb = load_workbook(f'{data[3]} TC Plant Inventory List.xlsx')
        #Select sheet to edit
        select_sheet = data[4]
        ws = wb[select_sheet]
        enter_data(data, ws)
        wb.save(f'{data[3]} TC Plant Inventory List.xlsx')
        more_data = input("Would you like to enter another strain? (Y/N): ")
        if more_data == "N":
            wb.save(f'{data[3]} TC Plant Inventory List.xlsx')
            break
