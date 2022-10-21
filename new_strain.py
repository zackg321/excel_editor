import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


### A program to create a new strain in the inventory list

def make_new_strain_ws(wb, ws, strain_name):
    #make a copy of template sheet
    new_strain_ws = wb.copy_worksheet(ws)
    new_strain_ws.title = strain_name
    wb.move_sheet(new_strain_ws, -1)

def get_data():
    #get customer name from user
    wb_name = input("Please enter customer name: ")

    #get strain name from user
    strain_name = input("Please enter strain name: ")

    #get number of explants from user
    num_explants = input("Please enter number of explants: ")

    #get date from user
    date = input("Please enter the date: ")

    data = [strain_name, num_explants, date, wb_name]
    return data


def fill_new_strain_ws(ws, data):
    #enter strain name into cell A1 on new strain sheet
    ws.cell(row=1, column=2).value = data[0]
    #enter date into cell D3 on new strain sheet
    ws.cell(row=3, column=4).value = data[2]
    #enter number of explants into cell E3 on new strain sheet
    ws.cell(row=3, column=5).value = data[1]
    #enter info from user into correct cells on the master sheet

def get_next_empty_row(wb):
    ws = wb["Master List"]
    rows = ws.iter_rows(min_row = 1, max_row = ws.max_row, min_col = 3, max_col = 3)
    for a in rows:
        if a[0].value == None:
            cell_row = a[0].row
            break
    return cell_row

def fill_master_list(wb, row, strain_name):
    #strain name
    wb["Master List"].cell(row = row, column = 3).value = f"='{strain_name}'!{get_column_letter(2)}1"
    #stage 1 explants
    wb["Master List"].cell(row = row, column = 4).value = f"='{strain_name}'!{get_column_letter(2)}2"
    #stage 1 date
    wb["Master List"].cell(row = row, column = 9).value = f"='{strain_name}'!{get_column_letter(2)}3"
    #stage 2 explants
    wb["Master List"].cell(row = row, column = 5).value = f"='{strain_name}'!{get_column_letter(2)}5"
    #stage 2 date
    wb["Master List"].cell(row = row, column = 11).value = f"='{strain_name}'!{get_column_letter(2)}6"
    #stage 2.5 explants
    wb["Master List"].cell(row = row, column = 6).value = f"='{strain_name}'!{get_column_letter(2)}11"
    #stage 2.5 date
    wb["Master List"].cell(row = row, column = 13).value = f"='{strain_name}'!{get_column_letter(2)}12"
    #stage 3 explants 
    wb["Master List"].cell(row = row, column = 7).value = f"='{strain_name}'!{get_column_letter(2)}8"
    #stage 3 date
    wb["Master List"].cell(row = row, column = 15).value = f"='{strain_name}'!{get_column_letter(2)}9"

    #sum of explants
    wb["Master List"].cell(row = row, column = 8).value = f"=SUM(D{row}:G{row})"
    #stage 1 ready?
    wb["Master List"].cell(row = row, column = 10).value = f'=IF(TODAY() - I{row} >= 21, "Yes", "No")'
    #stage 2 ready?
    wb["Master List"].cell(row = row, column = 12).value = f'=IF(TODAY() - K{row} >= 21, "Yes", "No")'
    #stage 2.5 ready?
    wb["Master List"].cell(row = row, column = 14).value = f'=IF(TODAY() - M{row} >= 21, "Yes", "No")'
    #stage 3 ready?
    wb["Master List"].cell(row = row, column = 16).value = f'=IF(TODAY() - O{row} >= 21, "Yes", "No")'

def run():
    data = get_data()
    wb_name = data[3]
    #load workbook
    wb = load_workbook(f"{wb_name} TC Plant Inventory List.xlsx")
    ws = wb["Template"]

    row = get_next_empty_row(wb)
    make_new_strain_ws(wb, ws, data[0])
    ws = wb[data[0]]
    fill_new_strain_ws(ws, data)
    fill_master_list(wb, row, data[0])
    wb.save(f"{wb_name} TC Plant Inventory List.xlsx")


